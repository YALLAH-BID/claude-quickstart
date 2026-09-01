#!/usr/bin/env python3
"""Automall daily used-vehicle stock pipeline.

Takes the two daily exports and produces two deliverables:

  1. Automall_Stock_Analysis_<date>.xlsx  - formula-driven analysis workbook
     (overview KPIs, brand/model months-of-stock, aging, aged-vehicle list,
     pricing flags, stocklist-vs-raw reconciliation, embedded data sheets).
  2. Stock_Raw_data_<date>_REFRESHED.xlsx - the raw-data workbook with the
     Automall rows of its UAE-STOCK sheet surgically corrected from the
     stocklist (direct XML edits: pivot tables, external links, formulas and
     all other sheets are preserved byte-for-byte).

Usage:
  python3 pipeline.py --stocklist STOCKLIST_28.08.2026.xlsx \
                      --rawdata "28.08.2026 (Stock Raw data).xlsx" \
                      --outdir out [--asof 28.08.2026]

After building, recalculate the analysis workbook (LibreOffice must have the
Calc component installed - `apt-get install -y libreoffice-calc`), e.g. with
the Claude xlsx skill's recalc.py. The refreshed raw-data workbook must NOT
be run through LibreOffice (it has external links); it is written with
fullCalcOnLoad so Excel recalculates it on open.

Input expectations (stable export formats):
  - Stocklist: single sheet, headers in row 1, VIN in column C
    ("Vehicle ID No."), the AFM.PUR.PRO.1F10 form layout (49 columns A..AW).
  - Raw data: sheets named "UAE-STOCK" (28 columns A..AB), "PIPELINE"
    (VIN in column I), and a trailing-3-months sold sheet whose name starts
    with "SOLD DATA (" (35 columns A..AI).
Only openpyxl is required (pip install openpyxl).
"""

import argparse
import json
import re
import sys
import zipfile
from collections import Counter
from datetime import date, datetime
from xml.sax.saxutils import escape

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------- helpers


def s(x):
    return str(x).strip() if x is not None else ""


def num(x):
    """Numeric view of a cell: numbers pass through, ''/'-'/None -> 0,
    numeric text parses, anything else -> None."""
    if isinstance(x, (int, float)):
        return x
    if x in (None, "", "-"):
        return 0.0
    try:
        return float(str(x).strip().replace(",", ""))
    except ValueError:
        return None


def vnum(x):
    return x if isinstance(x, (int, float)) else 0.0


MONTH_ORDER = [
    "January",
    "February",
    "March",
    "April",
    "May",
    "June",
    "July",
    "August",
    "September",
    "October",
    "November",
    "December",
]


# ----------------------------------------------------------------- loading


def load_inputs(stocklist_path, rawdata_path):
    wb = load_workbook(stocklist_path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    shdr = [" ".join(str(h).split()) if h is not None else "" for h in rows[0]]
    if "Vehicle ID No." not in shdr:
        raise SystemExit(
            f"Stocklist sheet {ws.title!r} has no 'Vehicle ID No.' header - "
            f"got {shdr[:6]}..."
        )
    sdata = [r for r in rows[1:] if r[2] not in (None, "")]

    wb = load_workbook(rawdata_path, read_only=True, data_only=True)
    names = wb.sheetnames
    if "UAE-STOCK" not in names or "PIPELINE" not in names:
        raise SystemExit(
            f"Raw-data file must contain 'UAE-STOCK' and 'PIPELINE' sheets; got {names}"
        )
    sold_names = [n for n in names if n.upper().startswith("SOLD DATA (")]
    if not sold_names:
        raise SystemExit(
            f"No trailing-3-months sold sheet (name starting 'SOLD DATA (') found; got {names}"
        )
    sold_name = sold_names[0]

    uae_rows = list(wb["UAE-STOCK"].iter_rows(values_only=True))
    uhdr = [" ".join(str(h).split()) if h is not None else "" for h in uae_rows[0]]
    uae = [r for r in uae_rows[1:] if r[1] not in (None, "")]

    pipe_rows = list(wb["PIPELINE"].iter_rows(values_only=True))
    phdr = [" ".join(str(h).split()) if h is not None else "" for h in pipe_rows[0]]
    pipe = [r for r in pipe_rows[1:] if r[8] not in (None, "")]

    sold_rows = list(wb[sold_name].iter_rows(values_only=True))
    sold = [r for r in sold_rows[1:] if r[0] not in (None, "")]
    wb.close()

    # months in chronological order (by earliest parseable billing date, else calendar)
    first_seen = {}
    for r in sold:
        m = s(r[17])
        if not m:
            continue
        d = None
        try:
            d = datetime.strptime(s(r[16]), "%d-%b-%y")
        except ValueError:
            pass
        cur = first_seen.get(m)
        if cur is None or (d is not None and (cur is None or d < cur)):
            first_seen[m] = d if d is not None else cur
    months = sorted(
        first_seen,
        key=lambda m: (
            first_seen[m] or datetime.max,
            MONTH_ORDER.index(m) if m in MONTH_ORDER else 99,
        ),
    )

    companies = [c for c, _ in Counter(s(r[25]) for r in uae if s(r[25])).most_common()]
    return {
        "shdr": shdr,
        "sdata": sdata,
        "uhdr": uhdr,
        "uae": uae,
        "phdr": phdr,
        "pipe": pipe,
        "sold": sold,
        "sold_name": sold_name,
        "months": months,
        "companies": companies,
    }


# ---------------------------------------------------------- reconciliation

RECON_FIELDS = [
    # (label, stk_idx, uae_idx, kind, tol_abs, tol_rel)
    ("KM", 12, 8, "num", 0.5, 0),
    ("Age (days)", 17, 9, "num", 0.5, 0),
    ("Datum / Purchase Price", 13, 11, "num", 0.5, 0.005),
    ("RSP incl VAT", 22, 12, "num", 0.5, 0.005),
    ("Availability", 31, 14, "str", 0, 0),
    ("Prep Cost", 15, 15, "num", 0.5, 0.005),
    ("Prep %", 16, 16, "pct", 0.0005, 0),
    ("SIV Cost Price", 14, 17, "num", 0.5, 0.005),
    ("GM", 23, 18, "num", 0.5, 0.005),
    ("GM %", 24, 19, "pct", 0.0005, 0),
    ("Buyer", 0, 22, "str", 0, 0),
    ("Vehicle Usage", 29, 13, "str", 0, 0),
]


def build_recon(data):
    sdata, uae = data["sdata"], data["uae"]
    uidx = {s(r[1]): r for r in uae}
    matched_rows = [r for r in sdata if s(r[2]) in uidx]

    raw_buyers = {s(uidx[s(r[2])][22]) for r in matched_rows}
    buyer_copydown = len(raw_buyers) == 1 and len(matched_rows) > 25

    rows, counts = [], Counter()
    for r in matched_rows:
        vin, make, model = s(r[2]), s(r[3]), s(r[4])
        u = uidx[vin]
        for label, si, ui, kind, ta, tr in RECON_FIELDS:
            a, b = r[si], u[ui]
            if kind in ("num", "pct"):
                an, bn = num(a), num(b)
                if an is None or bn is None:
                    if s(a) == s(b):
                        continue
                    d = (an or 0.0) - (bn or 0.0)
                else:
                    d = an - bn
                    if abs(d) <= ta or (tr and abs(d) <= tr * max(abs(an), abs(bn))):
                        continue
                cat, note = "Data", "Differs between the two files"
                if label == "RSP incl VAT" and an and bn and abs(bn - an * 1.05) < 1:
                    note = "Raw = stocklist x 1.05 (VAT added despite margin scheme)"
                if label == "Age (days)" and abs(d) <= 1.6:
                    cat, note = "Timing", "Raw file aged as of a different day"
                rows.append((vin, make, model, label, a, b, round(d, 2), cat, note))
            else:
                sa, sb = s(a), s(b)
                if sa.upper() == sb.upper():
                    continue
                cat, note = "Data", "Differs between the two files"
                if label == "Vehicle Usage" and {sa.upper(), sb.upper()} == {
                    "LCV",
                    "LV",
                }:
                    cat, note = (
                        "Convention",
                        "Same category, different code (LCV vs LV)",
                    )
                if label == "Buyer" and buyer_copydown:
                    note = (
                        "Raw buyer column holds a single value for every Automall row "
                        "(copy-down error)"
                    )
                if label == "Availability":
                    cat, note = (
                        "Convention",
                        f"Stocklist shows {sa!r}; raw file shows {sb!r}",
                    )
                rows.append((vin, make, model, label, sa, sb, "", cat, note))
            counts[label] += 1
    return {"rows": rows, "counts": counts, "matched": len(matched_rows)}


# --------------------------------------------------------- analysis workbook

F = "Arial"
NUMF, MON, PCT, DEC1, DATEFMT = "#,##0", "#,##0", "0.0%", "0.0", "DD-MMM-YY"
BUCKETS = [
    ("0-30", None, 30),
    ("31-60", 30, 60),
    ("61-90", 60, 90),
    ("91-120", 90, 120),
    ("121-150", 120, 150),
    ("150+", 150, None),
]


def build_analysis(data, recon, out_path, asof):
    sdata, uae, pipe, sold = data["sdata"], data["uae"], data["pipe"], data["sold"]
    shdr, uhdr, phdr = data["shdr"], data["uhdr"], data["phdr"]
    months, companies = data["months"], data["companies"]

    font_base = Font(name=F, size=10)
    font_title = Font(name=F, size=14, bold=True, color="1F3864")
    font_sub = Font(name=F, size=9, italic=True, color="595959")
    font_section = Font(name=F, size=11, bold=True, color="FFFFFF")
    font_hdr = Font(name=F, size=10, bold=True, color="FFFFFF")
    font_bold = Font(name=F, size=10, bold=True)
    font_note = Font(name=F, size=9, italic=True, color="595959")
    fill_section = PatternFill("solid", fgColor="1F3864")
    fill_hdr = PatternFill("solid", fgColor="4472C4")
    fill_total = PatternFill("solid", fgColor="D9E2F2")
    fill_warn = PatternFill("solid", fgColor="FCE4D6")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def sect(ws, row, text, width):
        c = ws.cell(row=row, column=1, value=text)
        c.font = font_section
        c.fill = fill_section
        for j in range(2, width + 1):
            ws.cell(row=row, column=j).fill = fill_section

    def header_row(ws, row, headers, start_col=1):
        for j, h in enumerate(headers, start=start_col):
            c = ws.cell(row=row, column=j, value=h)
            c.font = font_hdr
            c.fill = fill_hdr
            c.border = border
            c.alignment = Alignment(
                wrap_text=True, vertical="center", horizontal="center"
            )

    def put(ws, row, col, value, fmt=None, bold=False, bordered=True, align=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font = font_bold if bold else font_base
        if fmt is None and isinstance(value, (datetime, date)):
            fmt = DATEFMT
        if fmt:
            c.number_format = fmt
        if bordered:
            c.border = border
        if align:
            c.alignment = Alignment(horizontal=align)
        return c

    onfloor = [r for r in sdata if s(r[31])]
    n_pipe_in_stk = len(sdata) - len(onfloor)
    months_label = f"{months[0]}-{months[-1]}" if months else "last 3 months"
    n_months = max(len(months), 1)
    asof_long = asof.strftime("%d %B %Y").lstrip("0")
    asof_tag = asof.strftime("%d.%m.%Y")

    wb = Workbook()

    # ---- READ ME
    ws = wb.active
    ws.title = "READ ME"
    ws.sheet_properties.tabColor = "375623"
    ws.column_dimensions["A"].width = 118
    lines = [
        (f"AUTOMALL - USED VEHICLE STOCK ANALYSIS - {asof_long.upper()}", "title"),
        ("", None),
        ("Sources", "h"),
        (
            f"  1. Stocklist {asof_tag} ({len(sdata):,} Automall vehicles: {len(onfloor):,} on-floor "
            f"+ {n_pipe_in_stk:,} incoming pipeline).",
            None,
        ),
        (
            f"  2. Stock Raw data {asof_tag} - sheets UAE-STOCK ({len(uae):,} vehicles, all companies), "
            f"PIPELINE ({len(pipe):,}), {data['sold_name']} ({len(sold):,} invoices).",
            None,
        ),
        ("  All amounts are AED. UAE VAT rate 5%.", None),
        ("", None),
        ("Sheets", "h"),
        (
            "  OVERVIEW - network and Automall KPIs, sales run-rate, months of stock (MOS), aging, data quality.",
            None,
        ),
        (
            "  BRAND-MODEL MOS - stock, pipeline, 3-month sales and MOS for every brand-model combination.",
            None,
        ),
        (
            "  AGING - stock aging buckets by company and, for the Automall stocklist, by make.",
            None,
        ),
        (
            "  AGED VEHICLES - on-floor Automall vehicles aged over 90 days (provision-risk view).",
            None,
        ),
        (
            "  PRICING FLAGS - unpriced vehicles, negative or thin margins, price below cost, aged-not-ready stock.",
            None,
        ),
        (
            "  RECONCILIATION - field-by-field differences between the stocklist and the raw file's UAE-STOCK sheet.",
            None,
        ),
        (
            "  DATA_* sheets - the source data embedded as values; every analysis figure is a live formula over them.",
            None,
        ),
        ("", None),
        ("Method and assumptions", "h"),
        (
            f"  - MOS (months of stock) = units in stock / (units invoiced {months_label} / {n_months}), the same",
            None,
        ),
        (
            "    convention as the raw file's SUMMARY sheet, but matched on brand + model (not model name alone).",
            None,
        ),
        (
            "  - Aging buckets (0-30, 31-60, 61-90, 91-120, 121-150, 150+ days) are presentational, not a provision policy.",
            None,
        ),
        (
            "  - The stocklist's 'Age Provision Current/Next Month' columns carry no amounts, so AGED VEHICLES uses",
            None,
        ),
        ("    age > 90 days as the risk proxy instead.", None),
        (
            f"  - The {n_pipe_in_stk:,} stocklist vehicles without an availability status are the incoming pipeline;",
            None,
        ),
        ("    they are excluded from on-floor pricing/aging measures.", None),
        (
            "  - 'Margin scheme = YES' vehicles carry VAT on the dealer margin only, so their VAT-inclusive selling",
            None,
        ),
        ("    price equals the net price.", None),
        (
            "  - Data sheets hold the files exactly as uploaded; differences between them are catalogued in",
            None,
        ),
        (
            "    RECONCILIATION, and a separately delivered refreshed copy of the raw-data file applies the stocklist values.",
            None,
        ),
        ("", None),
        (
            f"Generated by stock_report/pipeline.py on {date.today().strftime('%d %B %Y').lstrip('0')}",
            "note",
        ),
    ]
    r = 1
    for text, kind in lines:
        c = ws.cell(row=r, column=1, value=text)
        c.font = {
            "title": font_title,
            "h": Font(name=F, size=11, bold=True, color="1F3864"),
            "note": font_note,
        }.get(kind, font_base)
        r += 1

    # ---- data sheets
    def data_sheet(name, headers, rows, fmts=None, tab="808080"):
        w = wb.create_sheet(name)
        w.sheet_properties.tabColor = tab
        header_row(w, 1, headers)
        for i, row in enumerate(rows, start=2):
            for j, v in enumerate(row, start=1):
                c = w.cell(row=i, column=j, value=v)
                c.font = font_base
                if fmts and headers[j - 1] in fmts:
                    c.number_format = fmts[headers[j - 1]]
                elif isinstance(v, (datetime, date)):
                    c.number_format = DATEFMT
        w.freeze_panes = "A2"
        w.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
        for j in range(1, len(headers) + 1):
            w.column_dimensions[get_column_letter(j)].width = 14
        return w

    stk_fmts = {
        "Mileage": NUMF,
        "Datum": MON,
        "SIV / COST PRICE": MON,
        "PREP": MON,
        "PREP%": PCT,
        "Age": "0",
        "DSR": MON,
        "Price": MON,
        "Selling Price (VAT Inclusive)": MON,
        "GM": MON,
        "GM%": PCT,
    }
    data_sheet("DATA_STOCKLIST", shdr, sdata, stk_fmts)
    uae_fmts = {
        "AGE": "0",
        "MOS MODEL LEVEL": DEC1,
        "Purchase Price / DATUM": MON,
        "RSP incl. VAT": MON,
        "Prep Cost": MON,
        "Prep %": PCT,
        "SIV Cost Price": MON,
        "GM": MON,
        "GM%": PCT,
        "Estimated Margin": MON,
        "Est GM %": PCT,
        "Avg. Provision": MON,
    }
    data_sheet("DATA_UAESTOCK", uhdr[:28], [tuple(r[:28]) for r in uae], uae_fmts)
    pipe_fmts = {
        "PREP": MON,
        "MILEAGE": NUMF,
        "DATUM": MON,
        "SUGGESTED NET SELLING PRICE": MON,
        "PRICE CHECK": MON,
    }
    data_sheet("DATA_PIPELINE", phdr[:40], [tuple(r[:40]) for r in pipe], pipe_fmts)
    sold_hdr = [
        "VIN",
        "Brand",
        "Model",
        "Grade",
        "Model Year",
        "KM",
        "Company",
        "Billing Date",
        "Month of Billing",
        "Net Revenue (AED)",
        "Gross Margin (AED)",
        "GM %",
    ]
    sold_rows = [
        (r[0], r[1], r[2], r[3], r[6], r[7], r[15], r[16], r[17], r[31], r[33], r[34])
        for r in sold
    ]
    data_sheet(
        "DATA_SOLD_3M",
        sold_hdr,
        sold_rows,
        {"Net Revenue (AED)": MON, "Gross Margin (AED)": MON, "GM %": PCT},
    )

    STK_LAST, UAE_LAST = len(sdata) + 1, len(uae) + 1
    PIPE_LAST, SOLD_LAST = len(pipe) + 1, len(sold) + 1

    def stk_rng(col):
        return f"DATA_STOCKLIST!${col}$2:${col}${STK_LAST}"

    def uae_rng(col):
        return f"DATA_UAESTOCK!${col}$2:${col}${UAE_LAST}"

    def pipe_rng(col):
        return f"DATA_PIPELINE!${col}$2:${col}${PIPE_LAST}"

    def sold_rng(col):
        return f"DATA_SOLD_3M!${col}$2:${col}${SOLD_LAST}"

    def bucket_crit(rng, lo, hi):
        parts = []
        if lo is not None:
            parts.append(f'{rng},">{lo}"')
        if hi is not None:
            parts.append(f'{rng},"<={hi}"')
        return ",".join(parts)

    # ---- field-order block used by OVERVIEW + RECONCILIATION
    def field_note(label):
        """Describe a field's differences from the data, not from a fixed script.

        The shape of these differences changes between exports (a column that is
        copy-down corrupt one day may simply be missing the next), so a hardcoded
        sentence goes stale and misreports. Everything below is measured.
        """
        sub = [x for x in recon["rows"] if x[3] == label]
        n = len(sub)
        if not n:
            return "No differences"
        if label == "Age (days)":
            return "Raw file computed age as of a different day"
        if label == "Vehicle Usage":
            return "Same category, different code (LCV in stocklist, LV in raw file)"
        if label == "Availability":
            return "Different status wording between the files"
        raw_blank = sum(1 for x in sub if x[5] in (None, "", "-"))
        if label == "RSP incl VAT":
            vat = sum(1 for x in sub if "VAT added" in x[8])
            parts = []
            if raw_blank:
                parts.append(f"Raw file has no price on {raw_blank} of {n} rows")
            if vat:
                parts.append(
                    f"{vat} are the margin-scheme VAT case (raw = stocklist x 1.05)"
                )
            return "; ".join(parts) if parts else "Differs between the files"
        seen = Counter(
            round(x[5], 4) if isinstance(x[5], (int, float)) else x[5] for x in sub
        )
        value, hits = seen.most_common(1)[0]
        if hits == n and n > 25:
            return f"Raw column holds the single value {value!r} on all {n} rows (copy-down error)"
        if raw_blank == n:
            return f"Raw file has no value on any of the {n} rows"
        if raw_blank > n / 2:
            return f"Raw file has no value on {raw_blank} of {n} rows"
        if raw_blank:
            return (
                f"Differs between the files; raw file has no value on "
                f"{raw_blank} of {n} rows"
            )
        return "Differs between the files"

    field_order = sorted(RECON_FIELDS, key=lambda f: -recon["counts"].get(f[0], 0))
    field_order = [f[0] for f in field_order]
    FIELD_ROW = {name: 12 + i for i, name in enumerate(field_order)}
    RC_DHDR = 12 + len(field_order) + 2

    # ---- OVERVIEW
    ov = wb.create_sheet("OVERVIEW", 1)
    ov.sheet_properties.tabColor = "1F3864"
    ov.sheet_view.showGridLines = False
    for col, wdt in {
        "A": 34,
        "B": 13,
        "C": 13,
        "D": 15,
        "E": 16,
        "F": 17,
        "G": 13,
        "H": 13,
        "I": 13,
    }.items():
        ov.column_dimensions[col].width = wdt
    ov["A1"] = "AUTOMALL - USED VEHICLE STOCK ANALYSIS"
    ov["A1"].font = font_title
    ov["A2"] = (
        f"As of {asof_long}  |  values in AED  |  sources: stocklist + stock raw data "
        f"{asof_tag}"
    )
    ov["A2"].font = font_sub

    r = 4
    sect(ov, r, "UAE NETWORK STOCK - ALL COMPANIES (UAE-STOCK sheet)", 9)
    r += 1
    header_row(
        ov,
        r,
        [
            "Company",
            "Units",
            "Available",
            "Customer Tagged",
            "SIV Value (AED)",
            "RSP Value (AED incl VAT)",
            "Avg Age (days)",
            f"Sold {months_label}",
            "MOS",
        ],
    )
    first = r + 1
    for i, comp in enumerate(companies):
        rr = r + 1 + i
        put(ov, rr, 1, comp)
        put(ov, rr, 2, f"=COUNTIFS({uae_rng('Z')},$A{rr})", NUMF)
        put(
            ov,
            rr,
            3,
            f'=COUNTIFS({uae_rng("Z")},$A{rr},{uae_rng("O")},"Available")',
            NUMF,
        )
        put(
            ov,
            rr,
            4,
            f'=COUNTIFS({uae_rng("Z")},$A{rr},{uae_rng("O")},"Customer Tagged")',
            NUMF,
        )
        put(ov, rr, 5, f"=SUMIFS({uae_rng('R')},{uae_rng('Z')},$A{rr})", MON)
        put(ov, rr, 6, f"=SUMIFS({uae_rng('M')},{uae_rng('Z')},$A{rr})", MON)
        put(
            ov,
            rr,
            7,
            f'=IFERROR(AVERAGEIFS({uae_rng("J")},{uae_rng("Z")},$A{rr}),"-")',
            "0",
        )
        put(ov, rr, 8, f"=COUNTIFS({sold_rng('G')},$A{rr})", NUMF)
        put(ov, rr, 9, f'=IFERROR($B{rr}/($H{rr}/{n_months}),"-")', DEC1)
    last = r + len(companies)
    rr = last + 1
    put(ov, rr, 1, "TOTAL", bold=True)
    for colI, colL, fmt in [
        (2, "B", NUMF),
        (3, "C", NUMF),
        (4, "D", NUMF),
        (5, "E", MON),
        (6, "F", MON),
    ]:
        put(
            ov, rr, colI, f"=SUM({colL}{first}:{colL}{last})", fmt, bold=True
        ).fill = fill_total
    put(
        ov, rr, 7, f'=IFERROR(AVERAGE({uae_rng("J")}),"-")', "0", bold=True
    ).fill = fill_total
    put(ov, rr, 8, f"=SUM(H{first}:H{last})", NUMF, bold=True).fill = fill_total
    put(
        ov, rr, 9, f'=IFERROR(B{rr}/(H{rr}/{n_months}),"-")', DEC1, bold=True
    ).fill = fill_total
    sold_companies = Counter(s(r_[15]) for r_ in sold)
    excluded = sum(c for comp, c in sold_companies.items() if comp not in companies)
    if excluded:
        ov.cell(
            row=rr + 1,
            column=1,
            value=f"Sold total on this table excludes {excluded} invoices booked to "
            "companies that have no stock rows.",
        ).font = font_note

    r = rr + 3
    sect(ov, r, "AUTOMALL STOCKLIST", 9)
    r += 1
    kpis = [
        ("Vehicles on stocklist", f"=COUNTA({stk_rng('C')})", NUMF, ""),
        (
            "  On-floor stock (has availability status)",
            f'=COUNTIFS({stk_rng("AF")},"<>")',
            NUMF,
            "",
        ),
        ("    - Available", f'=COUNTIFS({stk_rng("AF")},"Available")', NUMF, ""),
        (
            "    - Customer Tagged",
            f'=COUNTIFS({stk_rng("AF")},"Customer Tagged")',
            NUMF,
            "",
        ),
        (
            "    - Returns Available",
            f'=COUNTIFS({stk_rng("AF")},"Returns Available")',
            NUMF,
            "",
        ),
        (
            "  Incoming pipeline (no status yet)",
            f'=COUNTIFS({stk_rng("AF")},"")',
            NUMF,
            "matches PIPELINE sheet",
        ),
        (
            "SIV cost value - on-floor (AED)",
            f'=SUMIFS({stk_rng("O")},{stk_rng("AF")},"<>")',
            MON,
            "",
        ),
        (
            "RSP value incl VAT - on-floor (AED)",
            f'=SUMIFS({stk_rng("W")},{stk_rng("AF")},"<>")',
            MON,
            "",
        ),
        (
            "  Priced vehicles (net price > 0)",
            f'=COUNTIFS({stk_rng("AF")},"<>",{stk_rng("V")},">0")',
            NUMF,
            "",
        ),
        (
            "  Unpriced vehicles (net price = 0)",
            f'=COUNTIFS({stk_rng("AF")},"<>",{stk_rng("V")},0)',
            NUMF,
            "see PRICING FLAGS",
        ),
        (
            "Expected GM on priced stock (AED)",
            f'=SUMIFS({stk_rng("X")},{stk_rng("AF")},"<>",{stk_rng("V")},">0")',
            MON,
            "",
        ),
        (
            "Expected GM % (GM / net price, priced stock)",
            f'=IFERROR(SUMIFS({stk_rng("X")},{stk_rng("AF")},"<>",{stk_rng("V")},">0")'
            f'/SUMIFS({stk_rng("V")},{stk_rng("AF")},"<>",{stk_rng("V")},">0"),"-")',
            PCT,
            "",
        ),
        (
            "Average age - on-floor (days)",
            f'=IFERROR(AVERAGEIFS({stk_rng("R")},{stk_rng("AF")},"<>"),"-")',
            "0",
            "",
        ),
        (
            "Vehicles aged > 90 days (on-floor)",
            f'=COUNTIFS({stk_rng("AF")},"<>",{stk_rng("R")},">90")',
            NUMF,
            "see AGED VEHICLES",
        ),
        (
            "Vehicles aged > 120 days (on-floor)",
            f'=COUNTIFS({stk_rng("AF")},"<>",{stk_rng("R")},">120")',
            NUMF,
            "",
        ),
    ]
    for label, formula, fmt, note in kpis:
        put(ov, r, 1, label, bordered=False)
        put(ov, r, 2, formula, fmt, bold=True, bordered=False, align="right")
        if note:
            ov.cell(row=r, column=3, value=note).font = font_note
        r += 1

    r += 1
    sect(ov, r, f"SALES RUN-RATE - INVOICES {months_label.upper()}", 9)
    r += 1
    header_row(
        ov,
        r,
        [
            "Month",
            "Units - all companies",
            "Units - Automall",
            "Net Revenue (AED)",
            "Gross Margin (AED)",
            "GM %",
        ],
    )
    first = r + 1
    for i, m in enumerate(months):
        rr = r + 1 + i
        put(ov, rr, 1, m)
        put(ov, rr, 2, f"=COUNTIFS({sold_rng('I')},$A{rr})", NUMF)
        put(
            ov,
            rr,
            3,
            f'=COUNTIFS({sold_rng("I")},$A{rr},{sold_rng("G")},"Automall")',
            NUMF,
        )
        put(ov, rr, 4, f"=SUMIFS({sold_rng('J')},{sold_rng('I')},$A{rr})", MON)
        put(ov, rr, 5, f"=SUMIFS({sold_rng('K')},{sold_rng('I')},$A{rr})", MON)
        put(ov, rr, 6, f'=IFERROR(E{rr}/D{rr},"-")', PCT)
    last = r + len(months)
    rr = last + 1
    put(ov, rr, 1, f"Total {n_months} months", bold=True)
    for colI, colL, fmt in [
        (2, "B", NUMF),
        (3, "C", NUMF),
        (4, "D", MON),
        (5, "E", MON),
    ]:
        put(
            ov, rr, colI, f"=SUM({colL}{first}:{colL}{last})", fmt, bold=True
        ).fill = fill_total
    put(ov, rr, 6, f'=IFERROR(E{rr}/D{rr},"-")', PCT, bold=True).fill = fill_total
    rr += 1
    put(ov, rr, 1, "Average per month", bold=True)
    put(ov, rr, 2, f"=B{rr - 1}/{n_months}", DEC1, bold=True)
    put(ov, rr, 3, f"=C{rr - 1}/{n_months}", DEC1, bold=True)
    put(ov, rr, 4, f"=D{rr - 1}/{n_months}", MON, bold=True)
    put(ov, rr, 5, f"=E{rr - 1}/{n_months}", MON, bold=True)
    put(ov, rr, 6, "")
    avg_row = rr
    rr += 1
    put(ov, rr, 1, f"MOS - all companies ({len(uae):,} units in stock)", bordered=False)
    put(
        ov,
        rr,
        2,
        f'=IFERROR(COUNTA({uae_rng("B")})/B{avg_row},"-")',
        DEC1,
        bold=True,
        bordered=False,
        align="right",
    )
    rr += 1
    put(ov, rr, 1, "MOS - Automall (on-floor stock)", bordered=False)
    put(
        ov,
        rr,
        2,
        f'=IFERROR(COUNTIFS({stk_rng("AF")},"<>")/C{avg_row},"-")',
        DEC1,
        bold=True,
        bordered=False,
        align="right",
    )

    r = rr + 2
    sect(ov, r, "AGING SNAPSHOT (days in stock)", 9)
    r += 1
    header_row(ov, r, ["Scope"] + [b[0] for b in BUCKETS] + ["Total"])
    rr = r + 1
    put(ov, rr, 1, "UAE network (UAE-STOCK)")
    for j, (_, lo, hi) in enumerate(BUCKETS, start=2):
        put(ov, rr, j, f"=COUNTIFS({bucket_crit(uae_rng('J'), lo, hi)})", NUMF)
    put(ov, rr, 8, f"=SUM(B{rr}:G{rr})", NUMF, bold=True)
    rr += 1
    put(ov, rr, 1, "Automall stocklist (on-floor)")
    for j, (_, lo, hi) in enumerate(BUCKETS, start=2):
        put(
            ov,
            rr,
            j,
            f'=COUNTIFS({stk_rng("AF")},"<>",{bucket_crit(stk_rng("R"), lo, hi)})',
            NUMF,
        )
    put(ov, rr, 8, f"=SUM(B{rr}:G{rr})", NUMF, bold=True)
    ov.cell(
        row=rr + 1,
        column=1,
        value="Full company- and make-level aging is on the AGING sheet.",
    ).font = font_note

    r = rr + 3
    sect(
        ov,
        r,
        "DATA QUALITY - STOCKLIST vs RAW-DATA FILE (details on RECONCILIATION)",
        9,
    )
    r += 1
    dq = [
        ("Automall VINs matched between the two files", "=RECONCILIATION!B5", ""),
        ("Stocklist VINs only in PIPELINE (incoming)", "=RECONCILIATION!B6", ""),
        ("UAE-STOCK Automall rows missing from stocklist", "=RECONCILIATION!B7", ""),
        ("Field-level differences catalogued", "=RECONCILIATION!B8", ""),
    ]
    for name in field_order[:5]:
        dq.append(
            (
                f"  {name} differs",
                f"=RECONCILIATION!C{FIELD_ROW[name]}",
                field_note(name),
            )
        )
    for label, formula, note in dq:
        put(ov, r, 1, label, bordered=False)
        put(ov, r, 2, formula, NUMF, bold=True, bordered=False, align="right")
        if note:
            ov.cell(row=r, column=3, value=note).font = font_note
        r += 1

    # ---- BRAND-MODEL MOS
    bm = wb.create_sheet("BRAND-MODEL MOS", 2)
    bm.sheet_properties.tabColor = "1F3864"
    canon = {}
    for src, bi, mi in [(uae, 2, 3), (pipe, 11, 12), (sold, 1, 2)]:
        for row in src:
            key = (s(row[bi]).upper(), s(row[mi]).upper())
            if key not in canon and key != ("", ""):
                canon[key] = (s(row[bi]), s(row[mi]))
    pairs = sorted(canon.values(), key=lambda p: (p[0].upper(), p[1].upper()))
    bm["A1"] = "STOCK AND MONTHS-OF-STOCK BY BRAND AND MODEL - UAE NETWORK"
    bm["A1"].font = font_title
    bm["A2"] = (
        f"All companies. MOS = units in stock / (units invoiced {months_label} / {n_months}). "
        "Rows cover every brand-model seen in stock, pipeline or recent sales."
    )
    bm["A2"].font = font_sub
    HDR = 4
    header_row(
        bm,
        HDR,
        [
            "Brand",
            "Model",
            "Available",
            "Customer Tagged",
            "Stock Total",
            "SIV Value (AED)",
            "Pipeline",
            f"Sold {months_label}",
            "Avg Sold / Month",
            "MOS",
            "Note",
        ],
    )
    for i, (b, m) in enumerate(pairs):
        rr = HDR + 1 + i
        put(bm, rr, 1, b)
        put(bm, rr, 2, m)
        put(
            bm,
            rr,
            3,
            f'=COUNTIFS({uae_rng("C")},$A{rr},{uae_rng("D")},$B{rr},{uae_rng("O")},"Available")',
            NUMF,
        )
        put(
            bm,
            rr,
            4,
            f'=COUNTIFS({uae_rng("C")},$A{rr},{uae_rng("D")},$B{rr},{uae_rng("O")},"Customer Tagged")',
            NUMF,
        )
        put(bm, rr, 5, f"=COUNTIFS({uae_rng('C')},$A{rr},{uae_rng('D')},$B{rr})", NUMF)
        put(
            bm,
            rr,
            6,
            f"=SUMIFS({uae_rng('R')},{uae_rng('C')},$A{rr},{uae_rng('D')},$B{rr})",
            MON,
        )
        put(
            bm, rr, 7, f"=COUNTIFS({pipe_rng('L')},$A{rr},{pipe_rng('M')},$B{rr})", NUMF
        )
        put(
            bm, rr, 8, f"=COUNTIFS({sold_rng('B')},$A{rr},{sold_rng('C')},$B{rr})", NUMF
        )
        put(bm, rr, 9, f"=H{rr}/{n_months}", DEC1)
        put(bm, rr, 10, f'=IF(I{rr}=0,"-",E{rr}/I{rr})', DEC1)
        put(
            bm,
            rr,
            11,
            f'=IF(AND(E{rr}=0,H{rr}>0),"Out of stock - sold recently",'
            f'IF(AND(E{rr}>0,H{rr}=0),"No sales in last {n_months} months",'
            f'IF(N(J{rr})>6,"High MOS (>6 months)","")))',
        )
    first, last = HDR + 1, HDR + len(pairs)
    rr = last + 1
    put(bm, rr, 1, "TOTAL", bold=True)
    put(bm, rr, 2, "")
    for colI, colL, fmt in [
        (3, "C", NUMF),
        (4, "D", NUMF),
        (5, "E", NUMF),
        (6, "F", MON),
        (7, "G", NUMF),
        (8, "H", NUMF),
    ]:
        put(
            bm, rr, colI, f"=SUM({colL}{first}:{colL}{last})", fmt, bold=True
        ).fill = fill_total
    put(bm, rr, 9, f"=H{rr}/{n_months}", DEC1, bold=True).fill = fill_total
    put(bm, rr, 10, f'=IF(I{rr}=0,"-",E{rr}/I{rr})', DEC1, bold=True).fill = fill_total
    put(bm, rr, 11, "")
    bm.freeze_panes = f"A{HDR + 1}"
    bm.auto_filter.ref = f"A{HDR}:K{last}"
    for col, wdt in zip(
        "ABCDEFGHIJK", [18, 24, 11, 15, 11, 15, 10, 13, 14, 8, 30], strict=True
    ):
        bm.column_dimensions[col].width = wdt

    # ---- AGING
    ag = wb.create_sheet("AGING", 3)
    ag.sheet_properties.tabColor = "1F3864"
    ag.sheet_view.showGridLines = False
    ag["A1"] = "STOCK AGING"
    ag["A1"].font = font_title
    ag["A2"] = "Days in stock; values in AED. Buckets are presentational."
    ag["A2"].font = font_sub

    def aging_block(
        start, title, row_labels, count_formula, value_formula=None, extra=None
    ):
        rr0 = start
        sect(ag, rr0, title, 9)
        rr0 += 1
        header_row(
            ag,
            rr0,
            ["Scope"] + [b[0] for b in BUCKETS] + ["Total", extra or "> 90 days"],
        )
        firstr = rr0 + 1
        for i, lab in enumerate(row_labels):
            rr = rr0 + 1 + i
            put(ag, rr, 1, lab)
            for j, (_, lo, hi) in enumerate(BUCKETS, start=2):
                put(
                    ag,
                    rr,
                    j,
                    count_formula(rr, lo, hi),
                    NUMF if value_formula is None else MON,
                )
            put(
                ag,
                rr,
                8,
                f"=SUM(B{rr}:G{rr})",
                NUMF if value_formula is None else MON,
                bold=True,
            )
            put(
                ag,
                rr,
                9,
                value_formula(rr) if value_formula else count_formula(rr, 90, None),
                MON if value_formula else NUMF,
            )
        lastr = rr0 + len(row_labels)
        rr = lastr + 1
        put(ag, rr, 1, "TOTAL", bold=True)
        for j, colL in enumerate("BCDEFGHI", start=2):
            put(
                ag,
                rr,
                j,
                f"=SUM({colL}{firstr}:{colL}{lastr})",
                MON if (value_formula or colL == "I") else NUMF,
                bold=True,
            ).fill = fill_total
        return rr + 2

    r = aging_block(
        4,
        "UNITS BY COMPANY AND AGE BUCKET (UAE-STOCK)",
        companies,
        lambda rr, lo, hi: (
            f"=COUNTIFS({uae_rng('Z')},$A{rr},{bucket_crit(uae_rng('J'), lo, hi)})"
        ),
    )
    r = aging_block(
        r,
        "SIV VALUE (AED) BY COMPANY AND AGE BUCKET (UAE-STOCK)",
        companies,
        lambda rr, lo, hi: (
            f"=SUMIFS({uae_rng('R')},{uae_rng('Z')},$A{rr},{bucket_crit(uae_rng('J'), lo, hi)})"
        ),
        value_formula=lambda rr: (
            f'=SUMIFS({uae_rng("R")},{uae_rng("Z")},$A{rr},{uae_rng("J")},">90")'
        ),
        extra="> 90 days",
    )
    makes = sorted({s(row[3]) for row in onfloor})
    r = aging_block(
        r,
        "AUTOMALL STOCKLIST (ON-FLOOR) - UNITS BY MAKE AND AGE BUCKET",
        makes,
        lambda rr, lo, hi: (
            f'=COUNTIFS({stk_rng("D")},$A{rr},{stk_rng("AF")},"<>",{bucket_crit(stk_rng("R"), lo, hi)})'
        ),
        value_formula=lambda rr: (
            f'=SUMIFS({stk_rng("O")},{stk_rng("D")},$A{rr},{stk_rng("AF")},"<>",{stk_rng("R")},">90")'
        ),
        extra="SIV > 90d (AED)",
    )
    for col, wdt in zip("ABCDEFGHI", [22, 10, 10, 10, 10, 10, 10, 11, 16], strict=True):
        ag.column_dimensions[col].width = wdt

    # ---- AGED VEHICLES
    av = wb.create_sheet("AGED VEHICLES", 4)
    av.sheet_properties.tabColor = "C00000"
    aged = [
        row for row in onfloor if isinstance(row[17], (int, float)) and row[17] > 90
    ]
    aged.sort(key=lambda row: -row[17])
    av["A1"] = "ON-FLOOR VEHICLES AGED OVER 90 DAYS (PROVISION-RISK VIEW)"
    av["A1"].font = font_title
    av["A2"] = (
        "Extracted from DATA_STOCKLIST (availability status set, Age > 90 days), sorted "
        "oldest first. The stocklist's age-provision columns carry no amounts, so this "
        "age-based list is the risk view."
    )
    av["A2"].font = font_sub
    av["A2"].alignment = Alignment(wrap_text=True)
    av.merge_cells("A2:P2")
    av.row_dimensions[2].height = 30
    HDR = 4
    cols = [
        ("VIN", 2),
        ("Make", 3),
        ("Model", 4),
        ("Model Year", 11),
        ("Mileage", 12),
        ("Age (days)", 17),
        ("Datum (AED)", 13),
        ("Prep (AED)", 15),
        ("SIV Cost (AED)", 14),
        ("Net Price (AED)", 21),
        ("RSP incl VAT (AED)", 22),
        ("GM (AED)", 23),
        ("GM %", 24),
        ("Availability", 31),
        ("Prep Status", 32),
        ("Photo Status", 34),
        ("Buyer", 0),
        ("Vendor", 39),
    ]
    afmt = {
        "Mileage": NUMF,
        "Age (days)": "0",
        "Datum (AED)": MON,
        "Prep (AED)": MON,
        "SIV Cost (AED)": MON,
        "Net Price (AED)": MON,
        "RSP incl VAT (AED)": MON,
        "GM (AED)": MON,
        "GM %": PCT,
    }
    header_row(av, HDR, [c[0] for c in cols])
    for i, row in enumerate(aged):
        rr = HDR + 1 + i
        for j, (label, idx) in enumerate(cols, start=1):
            put(av, rr, j, row[idx], afmt.get(label))
    first, last = HDR + 1, HDR + len(aged)
    rr = last + 1
    put(av, rr, 1, f"TOTAL - {len(aged)} vehicles", bold=True)
    for j, colL in [(7, "G"), (8, "H"), (9, "I"), (10, "J"), (11, "K"), (12, "L")]:
        put(
            av, rr, j, f"=SUM({colL}{first}:{colL}{last})", MON, bold=True
        ).fill = fill_total
    av.freeze_panes = f"A{HDR + 1}"
    av.auto_filter.ref = f"A{HDR}:R{max(last, HDR + 1)}"
    for j, wdt in enumerate(
        [20, 13, 14, 10, 10, 10, 12, 11, 13, 13, 15, 11, 8, 15, 15, 12, 12, 26], start=1
    ):
        av.column_dimensions[get_column_letter(j)].width = wdt

    # ---- PRICING FLAGS
    pf = wb.create_sheet("PRICING FLAGS", 5)
    pf.sheet_properties.tabColor = "C00000"
    pf.sheet_view.showGridLines = False
    pf["A1"] = "PRICING AND READINESS FLAGS - AUTOMALL ON-FLOOR STOCK"
    pf["A1"].font = font_title
    pf["A2"] = (
        "On-floor = stocklist rows with an availability status (pipeline vehicles excluded)."
    )
    pf["A2"].font = font_sub
    r = 4
    sect(pf, r, "SUMMARY (live counts over DATA_STOCKLIST)", 8)
    r += 1
    flags_def = [
        (
            "F1",
            "Unpriced - net price = 0",
            f'=COUNTIFS({stk_rng("AF")},"<>",{stk_rng("V")},0)',
        ),
        (
            "F2",
            "Negative GM on a priced vehicle",
            f'=COUNTIFS({stk_rng("AF")},"<>",{stk_rng("V")},">0",{stk_rng("X")},"<0")',
        ),
        (
            "F3",
            "Thin margin - GM% between 0 and 5% (priced)",
            f'=COUNTIFS({stk_rng("AF")},"<>",{stk_rng("V")},">0",{stk_rng("Y")},">=0",{stk_rng("Y")},"<0.05")',
        ),
        (
            "F4",
            "Net price below SIV cost (priced)",
            f'=SUMPRODUCT(({stk_rng("AF")}<>"")*({stk_rng("V")}>0)*({stk_rng("V")}<{stk_rng("O")}))',
        ),
        (
            "F5",
            "Aged > 90 days and prep not READY",
            f'=COUNTIFS({stk_rng("AF")},"<>",{stk_rng("R")},">90",{stk_rng("AG")},"<>READY*")',
        ),
        (
            "F6",
            "Prep READY but no photo",
            f'=COUNTIFS({stk_rng("AF")},"<>",{stk_rng("AG")},"READY*",{stk_rng("AI")},"-")',
        ),
    ]
    header_row(pf, r, ["Flag", "Definition", "Vehicles"])
    for i, (code, label, formula) in enumerate(flags_def):
        rr = r + 1 + i
        put(pf, rr, 1, code)
        put(pf, rr, 2, label)
        put(pf, rr, 3, formula, NUMF, bold=True)
    r = r + len(flags_def) + 1
    pf.cell(
        row=r,
        column=1,
        value="A vehicle can raise several flags; the list below shows each flagged vehicle "
        "once with all of its flags.",
    ).font = font_note

    r += 2
    sect(pf, r, "FLAGGED VEHICLES (extracted from DATA_STOCKLIST)", 8)
    r += 1
    flagged = []
    for row in onfloor:
        fl = []
        V, X, Y = vnum(row[21]), vnum(row[23]), row[24]
        O_, R_, AG_ = vnum(row[14]), vnum(row[17]), s(row[32])
        if V == 0:
            fl.append("F1 unpriced")
        if V > 0 and X < 0:
            fl.append("F2 negative GM")
        if V > 0 and isinstance(Y, (int, float)) and 0 <= Y < 0.05:
            fl.append("F3 thin margin")
        if V > 0 and V < O_:
            fl.append("F4 price below cost")
        if R_ > 90 and not AG_.upper().startswith("READY"):
            fl.append("F5 aged, not ready")
        if AG_.upper().startswith("READY") and s(row[34]) == "-":
            fl.append("F6 ready, no photo")
        if fl:
            flagged.append((row, fl))
    flagged.sort(key=lambda t: (-len(t[1]), -vnum(t[0][17])))
    HDR = r
    fcols = [
        ("VIN", 2),
        ("Make", 3),
        ("Model", 4),
        ("Model Year", 11),
        ("Age (days)", 17),
        ("SIV Cost (AED)", 14),
        ("Net Price (AED)", 21),
        ("GM (AED)", 23),
        ("GM %", 24),
        ("Availability", 31),
        ("Prep Status", 32),
        ("Buyer", 0),
    ]
    ffmt = {
        "Age (days)": "0",
        "SIV Cost (AED)": MON,
        "Net Price (AED)": MON,
        "GM (AED)": MON,
        "GM %": PCT,
    }
    header_row(pf, HDR, [c[0] for c in fcols] + ["Flags"])
    for i, (row, fl) in enumerate(flagged):
        rr = HDR + 1 + i
        for j, (label, idx) in enumerate(fcols, start=1):
            put(pf, rr, j, row[idx], ffmt.get(label))
        c = put(pf, rr, len(fcols) + 1, "; ".join(fl))
        if any(f_.startswith(("F2", "F4")) for f_ in fl):
            c.fill = fill_warn
    pf.freeze_panes = f"A{HDR + 1}"
    pf.auto_filter.ref = f"A{HDR}:M{max(HDR + len(flagged), HDR + 1)}"
    for j, wdt in enumerate(
        [20, 13, 14, 10, 10, 13, 13, 11, 8, 15, 16, 12, 40], start=1
    ):
        pf.column_dimensions[get_column_letter(j)].width = wdt

    # ---- RECONCILIATION
    rc = wb.create_sheet("RECONCILIATION", 6)
    rc.sheet_properties.tabColor = "7F6000"
    rc["A1"] = (
        f"RECONCILIATION - STOCKLIST {asof_tag} vs RAW-DATA FILE (UAE-STOCK SHEET)"
    )
    rc["A1"].font = font_title
    rc["A2"] = (
        f"Comparison of the {recon['matched']:,} Automall VINs present in both files. "
        "Tolerance: 0.5 AED / 0.5% on amounts, 0.05pp on percentages; text compared "
        "case-insensitively. 'Convention' = same information, different code."
    )
    rc["A2"].font = font_sub
    put(rc, 4, 1, "VIN coverage", bold=True, bordered=False)
    cov = [
        (
            "Automall VINs in both files",
            f"=SUMPRODUCT(COUNTIF({uae_rng('B')},{stk_rng('C')}))",
        ),
        (
            "Stocklist VINs only in PIPELINE (incoming vehicles)",
            f"=COUNTA({stk_rng('C')})-B5",
        ),
        (
            "UAE-STOCK Automall rows missing from the stocklist",
            f'=COUNTIFS({uae_rng("Z")},"Automall")-B5',
        ),
        (
            "Field-level differences listed below",
            f"=COUNTA($A${RC_DHDR + 1}:$A$100000)",
        ),
    ]
    for i, (label, formula) in enumerate(cov):
        put(rc, 5 + i, 1, label, bordered=False)
        put(rc, 5 + i, 2, formula, NUMF, bold=True, bordered=False, align="right")
    put(rc, 10, 1, "Differences by field", bold=True, bordered=False)
    header_row(rc, 11, ["Field", "Category", "Rows"])
    cat_by_field = {}
    for row in recon["rows"]:
        cat_by_field.setdefault(row[3], row[7])
    for name in field_order:
        rr = FIELD_ROW[name]
        put(rc, rr, 1, name)
        put(rc, rr, 2, cat_by_field.get(name, "-"))
        put(rc, rr, 3, f"=COUNTIF($D${RC_DHDR + 1}:$D$100000,A{rr})", NUMF, bold=True)
        rc.cell(row=rr, column=4, value=field_note(name)).font = font_note
    header_row(
        rc,
        RC_DHDR,
        [
            "VIN",
            "Make",
            "Model",
            "Field",
            "Stocklist Value",
            "Raw-Data Value",
            "Delta",
            "Category",
            "Note",
        ],
    )
    for i, (vin, make, model, fieldn, a, b, d, cat, note) in enumerate(recon["rows"]):
        rr = RC_DHDR + 1 + i
        put(rc, rr, 1, vin)
        put(rc, rr, 2, make)
        put(rc, rr, 3, model)
        put(rc, rr, 4, fieldn)
        fmt = (
            PCT
            if fieldn in ("Prep %", "GM %")
            else MON
            if fieldn not in ("Age (days)", "KM")
            else NUMF
        )
        for j, v in [(5, a), (6, b), (7, d)]:
            vv = round(v, 4) if isinstance(v, float) else v
            put(
                rc,
                rr,
                j,
                vv if vv != "" else None,
                fmt if isinstance(vv, (int, float)) else None,
            )
        put(rc, rr, 8, cat)
        put(rc, rr, 9, note)
    rc.freeze_panes = f"A{RC_DHDR + 1}"
    rc.auto_filter.ref = f"A{RC_DHDR}:I{RC_DHDR + max(len(recon['rows']), 1)}"
    for j, wdt in enumerate([20, 12, 14, 20, 16, 16, 12, 11, 52], start=1):
        rc.column_dimensions[get_column_letter(j)].width = wdt

    wb.save(out_path)
    return {
        "pairs": len(pairs),
        "aged": len(aged),
        "flagged": len(flagged),
        "onfloor": len(onfloor),
        "months": months,
    }


# ----------------------------------------------------------- refresh output

REFRESH_TARGETS = [
    # (UAE-STOCK column letter, stocklist idx, uae idx, kind)
    ("I", 12, 8, "kmstr"),  # KM (stored as text in the raw file)
    ("J", 17, 9, "num"),  # AGE
    ("L", 13, 11, "num"),  # Purchase price / DATUM
    ("M", 22, 12, "num"),  # RSP incl VAT
    ("P", 15, 15, "num"),  # Prep Cost
    ("Q", 16, 16, "num"),  # Prep %
    ("S", 23, 18, "num"),  # GM
    ("T", 24, 19, "num"),  # GM %
    ("W", 0, 22, "str"),  # Buyer
]


def fmt_xml_num(x):
    if isinstance(x, float) and x == int(x) and abs(x) < 1e15:
        return str(int(x))
    return repr(x) if isinstance(x, float) else str(x)


def build_refresh(rawdata_path, data, out_path):
    sdata, uae = data["sdata"], data["uae"]
    uidx = {s(r[1]): (i + 2, r) for i, r in enumerate(uae)}

    edits, per_col = {}, Counter()
    for r in sdata:
        vin = s(r[2])
        if vin not in uidx:
            continue
        rowno, u = uidx[vin]
        for colL, si, ui, kind in REFRESH_TARGETS:
            a, b = r[si], u[ui]
            if kind in ("num", "kmstr"):
                an, bn = num(a), num(b)
                if an is None or (bn is not None and abs(an - bn) <= 0.005):
                    continue
                val = fmt_xml_num(an)
                edits.setdefault(rowno, {})[colL] = (
                    "str" if kind == "kmstr" else "num",
                    val,
                )
            else:
                sa, sb = s(a), s(b)
                if sa == sb or sa == "":
                    continue
                edits.setdefault(rowno, {})[colL] = ("str", sa)
            per_col[colL] += 1

    zin = zipfile.ZipFile(rawdata_path)
    wbxml = zin.read("xl/workbook.xml").decode("utf-8")
    rels = zin.read("xl/_rels/workbook.xml.rels").decode("utf-8")
    relmap = dict(re.findall(r'Id="(rId\d+)"[^>]*Target="([^"]+)"', rels))
    m = re.search(r'<sheet name="UAE-STOCK"[^>]*r:id="(rId\d+)"', wbxml)
    if not m:
        raise SystemExit("UAE-STOCK sheet not found in workbook.xml")
    sheet_part = "xl/" + relmap[m.group(1)].lstrip("/")
    sheet = zin.read(sheet_part).decode("utf-8")

    cell_re = re.compile(r'<c r="([A-Z]+)(\d+)"[^>]*?(?:/>|>.*?</c>)', re.S)
    srow_re = re.compile(r'(<row r="(\d+)"[^>]*?)(/>|>(.*?)</row>)', re.S)
    s_attr_re = re.compile(r'\ss="(\d+)"')

    col_style = {}
    row2 = re.search(r'<row r="2"[^>]*>(.*?)</row>', sheet, re.S)
    if row2:
        for cm in cell_re.finditer(row2.group(1)):
            sm = s_attr_re.search(cm.group(0).split(">", 1)[0])
            if sm:
                col_style.setdefault(cm.group(1), sm.group(1))

    def col_n(letters):
        v = 0
        for ch in letters:
            v = v * 26 + (ord(ch) - 64)
        return v

    def make_cell(colL, rowno, kind, value, style):
        s_part = f' s="{style}"' if style else ""
        if kind == "num":
            return f'<c r="{colL}{rowno}"{s_part}><v>{value}</v></c>'
        return (
            f'<c r="{colL}{rowno}"{s_part} t="inlineStr">'
            f'<is><t xml:space="preserve">{escape(value)}</t></is></c>'
        )

    stats = {"replaced": 0, "inserted": 0}

    def fix_row(mm):
        rowno = int(mm.group(2))
        if rowno not in edits:
            return mm.group(0)
        row_edits = dict(edits[rowno])
        opener, closer = mm.group(1), mm.group(3)
        inner = mm.group(4) if closer.startswith(">") else ""
        out_cells = []
        for cm in cell_re.finditer(inner):
            colL = cm.group(1)
            if colL in row_edits:
                kind, value = row_edits.pop(colL)
                sm = s_attr_re.search(cm.group(0).split(">", 1)[0])
                style = sm.group(1) if sm else col_style.get(colL)
                out_cells.append(
                    (col_n(colL), make_cell(colL, rowno, kind, value, style))
                )
                stats["replaced"] += 1
            else:
                out_cells.append((col_n(colL), cm.group(0)))
        for colL, (kind, value) in row_edits.items():
            out_cells.append(
                (col_n(colL), make_cell(colL, rowno, kind, value, col_style.get(colL)))
            )
            stats["inserted"] += 1
        out_cells.sort(key=lambda t: t[0])
        return opener + ">" + "".join(x for _, x in out_cells) + "</row>"

    new_sheet = srow_re.sub(fix_row, sheet)
    total = sum(len(v) for v in edits.values())
    if stats["replaced"] + stats["inserted"] != total:
        raise SystemExit(
            f"refresh edit mismatch: planned {total}, "
            f"applied {stats['replaced'] + stats['inserted']}"
        )

    if "<calcPr" in wbxml:
        if "fullCalcOnLoad" not in wbxml:
            new_wb = re.sub(r"<calcPr\b", '<calcPr fullCalcOnLoad="1"', wbxml, count=1)
        else:
            new_wb = wbxml
    else:
        new_wb = wbxml.replace("</sheets>", '</sheets><calcPr fullCalcOnLoad="1"/>', 1)

    with zipfile.ZipFile(out_path, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            payload = zin.read(item.filename)
            if item.filename == sheet_part:
                payload = new_sheet.encode("utf-8")
            elif item.filename == "xl/workbook.xml":
                payload = new_wb.encode("utf-8")
            zout.writestr(item, payload)
    zin.close()
    return {"cells": total, "per_column": dict(per_col), "rows": len(edits)}


# --------------------------------------------------------------- verification


def verify(data, recon, analysis_path, refresh_path, rawdata_path):
    problems = []
    sdata, uae = data["sdata"], data["uae"]

    # refreshed file: zip integrity, part inventory, residual diffs
    z = zipfile.ZipFile(refresh_path)
    if z.testzip() is not None:
        problems.append("refreshed workbook: zip CRC failure")
    if set(z.namelist()) != set(zipfile.ZipFile(rawdata_path).namelist()):
        problems.append("refreshed workbook: part inventory changed")
    z.close()
    wb = load_workbook(refresh_path, read_only=True, data_only=True)
    newu = {
        s(r[1]): r
        for r in wb["UAE-STOCK"].iter_rows(min_row=2, values_only=True)
        if r[1] not in (None, "")
    }
    wb.close()
    residual = 0
    for r in sdata:
        vin = s(r[2])
        if vin not in newu:
            continue
        u = newu[vin]
        for _colL, si, ui, kind in REFRESH_TARGETS:
            a, b = num(r[si]), num(u[ui])
            if kind == "str":
                if s(r[si]) and s(r[si]) != s(u[ui]):
                    residual += 1
            elif a is not None and (b is None or abs(a - b) > 0.005):
                residual += 1
    if residual:
        problems.append(
            f"refreshed workbook: {residual} residual differences vs stocklist"
        )
    stk_vins = {s(r[2]) for r in sdata}
    changed = 0
    oldu = {s(r[1]): r for r in uae}
    for vin, u_new in newu.items():
        if vin in stk_vins:
            continue
        if any(
            str(a) != str(b) for a, b in zip(oldu[vin][:28], u_new[:28], strict=False)
        ):
            changed += 1
    if changed:
        problems.append(f"refreshed workbook: {changed} non-Automall rows changed")

    # analysis workbook: structural checks (values need a recalc pass first)
    wb = load_workbook(analysis_path, read_only=True, data_only=False)
    expected = {
        "READ ME",
        "OVERVIEW",
        "BRAND-MODEL MOS",
        "AGING",
        "AGED VEHICLES",
        "PRICING FLAGS",
        "RECONCILIATION",
        "DATA_STOCKLIST",
        "DATA_UAESTOCK",
        "DATA_PIPELINE",
        "DATA_SOLD_3M",
    }
    missing = expected - set(wb.sheetnames)
    if missing:
        problems.append(f"analysis workbook: missing sheets {sorted(missing)}")
    nstk = sum(
        1
        for r in wb["DATA_STOCKLIST"].iter_rows(min_row=2, values_only=True)
        if r[2] not in (None, "")
    )
    if nstk != len(sdata):
        problems.append(
            f"analysis workbook: DATA_STOCKLIST has {nstk} rows, expected {len(sdata)}"
        )
    wb.close()
    return problems


# ----------------------------------------------------------------------- main


def main(argv=None):
    ap = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    ap.add_argument("--stocklist", required=True)
    ap.add_argument("--rawdata", required=True)
    ap.add_argument("--outdir", default=".")
    ap.add_argument(
        "--asof",
        help="DD.MM.YYYY; default: parsed from the stocklist filename, else today",
    )
    args = ap.parse_args(argv)

    if args.asof:
        asof = datetime.strptime(args.asof, "%d.%m.%Y").date()
    else:
        m = re.search(r"(\d{2}\.\d{2}\.\d{4})", args.stocklist)
        asof = datetime.strptime(m.group(1), "%d.%m.%Y").date() if m else date.today()
    tag = asof.strftime("%d.%m.%Y")

    data = load_inputs(args.stocklist, args.rawdata)
    recon = build_recon(data)

    import os

    os.makedirs(args.outdir, exist_ok=True)
    analysis_path = os.path.join(args.outdir, f"Automall_Stock_Analysis_{tag}.xlsx")
    refresh_path = os.path.join(args.outdir, f"Stock_Raw_data_{tag}_REFRESHED.xlsx")

    a_stats = build_analysis(data, recon, analysis_path, asof)
    r_stats = build_refresh(args.rawdata, data, refresh_path)
    problems = verify(data, recon, analysis_path, refresh_path, args.rawdata)

    summary = {
        "asof": tag,
        "stocklist_rows": len(data["sdata"]),
        "on_floor": a_stats["onfloor"],
        "uae_stock_rows": len(data["uae"]),
        "pipeline_rows": len(data["pipe"]),
        "sold_rows": len(data["sold"]),
        "months": a_stats["months"],
        "recon_matched": recon["matched"],
        "recon_differences": len(recon["rows"]),
        "recon_by_field": dict(recon["counts"]),
        "brand_model_pairs": a_stats["pairs"],
        "aged_over_90": a_stats["aged"],
        "flagged_vehicles": a_stats["flagged"],
        "refresh_cells": r_stats["cells"],
        "refresh_by_column": r_stats["per_column"],
        "outputs": {"analysis": analysis_path, "refreshed_rawdata": refresh_path},
        "verification_problems": problems,
    }
    print(json.dumps(summary, indent=2, default=str))
    if problems:
        print("\nVERIFICATION FAILED - do not deliver these outputs.", file=sys.stderr)
        return 1
    print(
        "\nNext: recalculate the ANALYSIS workbook (LibreOffice with Calc installed), "
        "e.g. the xlsx skill's recalc.py. Do NOT recalculate the refreshed raw-data "
        "workbook - it carries external links and recalculates itself in Excel on open.",
        file=sys.stderr,
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
